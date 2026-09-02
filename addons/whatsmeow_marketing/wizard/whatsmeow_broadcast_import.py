import base64
import io
import logging

from odoo import fields, models
from odoo.exceptions import UserError

from ..models.whatsmeow_broadcast_contact import phone_digits, phone_tail

_logger = logging.getLogger(__name__)

# A number with fewer digits than this is a typo, a note ("N/A", "-") or a
# local extension — never something WhatsApp can deliver to. Importing it would
# put a contact in the audience that every campaign then reports as failed.
MIN_PHONE_DIGITS = 7

# A guard, not a policy: openpyxl streams the rows, but a runaway file would
# still build a very large transaction. Well above any hand-maintained list.
MAX_ROWS = 20000

# Header aliases, lower-cased. The template's own wording comes first; the rest
# is what people actually name these columns when they rebuild the sheet by
# hand. Unrecognised headers fall back to column order (list, name, phone).
COLUMNS = {
    "list": ("broadcast name", "broadcast list", "broadcast", "list", "list name"),
    "name": ("name", "contact name", "contact", "full name"),
    "phone": ("phone", "phone number", "mobile", "number", "whatsapp", "whatsapp number"),
}


def _cell_text(value):
    """A spreadsheet cell as the string the user meant.

    Excel stores an unquoted number as a float, so a phone typed without its
    '+' arrives as `919746707744.0` — `str()` on that would import a contact
    nobody can message.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


class WhatsmeowBroadcastImport(models.TransientModel):
    """Build broadcast lists and their contacts from one spreadsheet.

    Deliberately thin, like the composer: it parses rows and calls the ordinary
    models, so ownership, the per-owner phone uniqueness and the opt-out flag
    all keep behaving exactly as they do when a contact is typed in by hand.

    An import never *undoes* a wish: a contact who opted out and appears in the
    file again stays opted out, and is only added to the new list.
    """
    _name = "whatsmeow.broadcast.import"
    _description = "Import Broadcast Lists"

    state = fields.Selection(
        [("upload", "Upload"), ("done", "Imported")],
        default="upload", required=True,
    )
    file = fields.Binary(string="Excel File", attachment=False)
    filename = fields.Char(string="File Name")

    default_list_id = fields.Many2one(
        "whatsmeow.broadcast.list", string="Default List",
        domain=lambda self: [("user_id", "=", self.env.uid)],
        help="Used for rows whose 'Broadcast Name' cell is empty and that have "
             "no named row above them. Leave it empty to require a name in the "
             "sheet.",
    )
    update_names = fields.Boolean(
        string="Update Existing Names", default=True,
        help="A number already in your contacts keeps its record — this decides "
             "whether the name in the file overwrites the stored one. The "
             "contact is added to the list either way.",
    )

    result_summary = fields.Text(string="Result", readonly=True)
    skipped_summary = fields.Text(string="Skipped Rows", readonly=True)
    imported_list_ids = fields.Many2many(
        "whatsmeow.broadcast.list", string="Lists", readonly=True,
    )

    # -- parsing --------------------------------------------------------------
    def _read_rows(self):
        """The sheet as `(row_number, {col: text})` dicts, header row dropped."""
        self.ensure_one()
        if not self.file:
            raise UserError(self.env._("Choose a file to import first."))
        name = (self.filename or "").lower()
        if name and not name.endswith((".xlsx", ".xlsm")):
            raise UserError(self.env._(
                "Only Excel files (.xlsx) are supported. Download the template "
                "below, fill it in and upload it back."))
        try:
            from openpyxl import load_workbook
            book = load_workbook(
                io.BytesIO(base64.b64decode(self.file)), read_only=True, data_only=True)
        except UserError:
            raise
        except Exception as exc:
            _logger.info("Broadcast import: unreadable workbook (%s)", exc)
            raise UserError(self.env._(
                "That file could not be read as an Excel workbook. Download the "
                "template below and fill it in."))

        sheet = book["Contacts"] if "Contacts" in book.sheetnames else book.worksheets[0]

        header = None
        header_row = 0
        records = []
        # Rows are numbered as Excel numbers them — 1-based, blank leading rows
        # included — so "Row 214" in the report is the row the user can click on.
        for index, raw in enumerate(sheet.iter_rows(values_only=True), start=1):
            cells = [_cell_text(cell) for cell in raw]
            if header is None:
                if any(cells):
                    header, header_row = cells, index
                continue
            if any(cells):
                records.append((index, cells))
                if len(records) > MAX_ROWS:
                    raise UserError(self.env._(
                        "That file has more than %s rows. Split it and import it "
                        "in parts.", MAX_ROWS))
        book.close()
        if header is None:
            raise UserError(self.env._("That file is empty."))

        mapping = {}
        labels = [cell.lower() for cell in header]
        for key, aliases in COLUMNS.items():
            for index, label in enumerate(labels):
                if label in aliases:
                    mapping[key] = index
                    break
        if "phone" not in mapping or "name" not in mapping:
            # No usable header: fall back to the template's column order rather
            # than refusing a sheet that is right but simply unlabelled — and
            # keep that first row, since it holds data in that case.
            mapping = {"list": 0, "name": 1, "phone": 2}
            records.insert(0, (header_row, header))

        return [
            (number, {
                key: (cells[index] if index < len(cells) else "")
                for key, index in mapping.items()
            })
            for number, cells in records
        ]

    # -- import ---------------------------------------------------------------
    def action_import(self):
        self.ensure_one()
        rows = self._read_rows()
        if not rows:
            raise UserError(self.env._("That file has no contact rows."))

        Contact = self.env["whatsmeow.broadcast.contact"]
        List = self.env["whatsmeow.broadcast.list"]
        uid = self.env.uid

        lists = {}          # lower-cased name -> record
        touched = List
        created_lists = 0
        created = updated = subscribed = 0
        seen_tails = {}     # tail -> row number, dedup within the file
        skipped = []
        current_name = ""

        for number, values in rows:
            if values.get("list"):
                current_name = values["list"]
            # A blank 'Broadcast Name' means "same list as the row above" —
            # that is how the template reads, and how anyone fills it in.
            target = None
            if current_name:
                key = current_name.lower()
                target = lists.get(key)
                if not target:
                    target = List.search(
                        [("name", "=ilike", current_name), ("user_id", "=", uid)], limit=1)
                    if not target:
                        target = List.create({"name": current_name, "user_id": uid})
                        created_lists += 1
                    lists[key] = target
            elif self.default_list_id:
                target = self.default_list_id
            if not target:
                skipped.append((number, self.env._("no broadcast list named")))
                continue
            touched |= target

            phone = values.get("phone", "")
            digits = phone_digits(phone)
            if len(digits) < MIN_PHONE_DIGITS:
                skipped.append((number, self.env._(
                    "'%s' is not a usable phone number", phone) if phone
                    else self.env._("no phone number")))
                continue
            tail = phone_tail(phone)
            if tail in seen_tails:
                skipped.append((number, self.env._(
                    "same number as row %s", seen_tails[tail])))
                continue
            seen_tails[tail] = number

            # A numbers-only list is a real thing to import; dropping the row
            # because nobody typed a name would lose the contact entirely.
            name = values.get("name") or phone

            # Archived too: the per-owner uniqueness index only covers active
            # rows, so ignoring an archived contact would build a second record
            # for the same person and hide their campaign history.
            contact = Contact.with_context(active_test=False).search(
                [("user_id", "=", uid), ("phone_tail", "=", tail)], limit=1)
            if contact:
                vals = {}
                if not contact.active:
                    vals["active"] = True
                if self.update_names and name != contact.name:
                    vals["name"] = name
                if target not in contact.list_ids:
                    vals["list_ids"] = [(4, target.id)]
                    subscribed += 1
                if vals:
                    contact.write(vals)
                updated += 1
            else:
                Contact.create({
                    "name": name,
                    "phone": phone,
                    "user_id": uid,
                    "list_ids": [(4, target.id)],
                })
                created += 1
                subscribed += 1

        summary = [
            self.env._("%s contacts created", created),
            self.env._("%s existing contacts reused", updated),
            self.env._("%s list memberships added", subscribed),
            self.env._("%s broadcast lists created", created_lists),
        ]
        if skipped:
            summary.append(self.env._("%s rows skipped", len(skipped)))

        self.write({
            "state": "done",
            "result_summary": "\n".join(summary),
            # Long enough to be useful, short enough to stay readable; the count
            # above is what says how much was dropped.
            "skipped_summary": "\n".join(
                self.env._("Row %(row)s: %(reason)s", row=number, reason=reason)
                for number, reason in skipped[:50]
            ) or False,
            "imported_list_ids": [(6, 0, touched.ids)],
        })
        return {
            "type": "ir.actions.act_window",
            "res_model": self._name,
            "res_id": self.id,
            "view_mode": "form",
            "view_id": self.env.ref(
                "whatsmeow_marketing.whatsmeow_broadcast_import_view_form").id,
            "target": "new",
        }

    def action_open_lists(self):
        self.ensure_one()
        return {
            "type": "ir.actions.act_window",
            "name": self.env._("Broadcast Lists"),
            "res_model": "whatsmeow.broadcast.list",
            "view_mode": "list,form",
            "domain": [("id", "in", self.imported_list_ids.ids)],
        }
